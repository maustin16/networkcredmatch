def sshOpen(hostname,username,password):
    #Opens a new ssh session to the requested hostname with the requested credentials
    print("Attempting to SSH to " + str(hostname) + " with the username " + username)
    ssh_client = paramiko.SSHClient()
    ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh_client.connect(hostname=hostname,username=username,password=password)
    return ssh_client

def sshClose(ssh_client):
    #Closes a passed ssh session
    print("Closing ssh client")
    ssh_client.close()
    return

def getCreds():
    #Opens the excel file and loads credentials
    wb = load_workbook('scriptworkbook.xlsx')
    ws = wb["creds"]
    creds = {}
    usernames = []
    passwords = []
    for row in range(1, (ws.max_row+1)):
        #For each row in the credentials section this will add the username to the list of usernames and the password to
        # the list of passwords
        usernames.append(ws.cell(row=row, column=1).value)
        passwords.append(ws.cell(row=row, column=2).value)
    #One dictionary for both usernames and passwords with the appropriate key
    creds["usernames"] = usernames
    creds["passwords"] = passwords
    return creds

def findDevices():
    import ipaddress
    wb = load_workbook('scriptworkbook.xlsx')
    ws = wb["subnets"]
    subnets = []
    devices = []
    for row in range(1, (ws.max_row+1)):
        subnets.append(ipaddress.ip_network(ws.cell(row=row, column=1).value))

    for network in subnets:
        print("Sweeping " + str(network))
        for host in network.hosts():
            host = str(host)
            resp = ping(host)
            if resp != None or False:
                print(str(host) + " is alive")
                devices.append(host)
            else:
                print(str(host) + " is dead")
    return devices

def testCreds(creds, devices):
    numberOfCreds = len(creds["usernames"])
    discovered = {}
    for device in devices:
        i = 0
        while i < numberOfCreds:
            username = creds["usernames"][i]
            password = creds["passwords"][i]
            print("Attempting to SSH to " + str(device) + " with the username " + username)
            try:
                ssh_client = sshOpen(device,username,password)
                print(str(device) + " authenticated with username " + username)
                sshClose(ssh_client)
                discovered[device] = i + 1
                break
            except Exception as e:
                print("Could not open SSH session to " + str(device) + " with username " + username + ". Reason: " + str(e))
            i = i + 1
    print(discovered)
    return discovered

def recordDiscovered(discovered):
    wb = load_workbook('scriptworkbook.xlsx')
    ws = wb["discovered"]
    row = 1
    for device in discovered:
        ws.cell(row=row,column=1,value=device)
        ws.cell(row=row,column=2,value=discovered[device])
        row = row + 1
    wb.save('scriptworkbook.xlsx')
    print("Saved Discovered credentials")

if __name__ == '__main__':
    import paramiko
    from ping3 import ping
    from openpyxl import load_workbook

    creds = getCreds()
    devices = findDevices()
    discovered = testCreds(creds,devices)
    recordDiscovered(discovered)
